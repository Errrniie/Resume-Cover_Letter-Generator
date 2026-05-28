#!/usr/bin/env python3
"""
Excel-backed application tracker at Data/applications.xlsx.
"""

from __future__ import annotations

import logging
import threading
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook

DATA_DIR_NAME = "Data"
TRACKER_FILENAME = "applications.xlsx"

HEADERS: tuple[str, ...] = (
    "date",
    "url",
    "website",
    "job_sequence",
    "company",
    "position",
    "resume",
    "cover_letter",
    "updated",
)

_lock = threading.Lock()
_log = logging.getLogger(__name__)


def project_root() -> Path:
    root = Path(__file__).resolve().parent
    return root.parent if root.name == "Modules" else root


@dataclass(frozen=True)
class ApplicationRecord:
    """One row from the application log."""

    row_number: int
    date: str
    url: str
    website: str
    job_sequence: int | None
    company: str
    position: str
    resume: str
    cover_letter: str
    updated: str


def tracker_path(*, base: Path | None = None) -> Path:
    return (base or project_root()) / DATA_DIR_NAME / TRACKER_FILENAME


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _website_from_url(url: str) -> str:
    parsed = urlparse(url.strip())
    return parsed.hostname or parsed.netloc or ""


def ensure_workbook(*, base: Path | None = None) -> None:
    """Create Data/applications.xlsx and header row if missing."""
    path = tracker_path(base=base)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.is_file():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "applications"
    for column, header in enumerate(HEADERS, start=1):
        sheet.cell(row=1, column=column, value=header)
    workbook.save(path)
    workbook.close()


def _header_map(sheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column in range(1, len(HEADERS) + 1):
        name = _cell_str(sheet.cell(row=1, column=column).value)
        if name:
            mapping[name] = column
    return mapping


def _row_to_record(sheet, row_number: int, columns: dict[str, int]) -> ApplicationRecord:
    def value(name: str) -> str:
        col = columns.get(name)
        if col is None:
            return ""
        return _cell_str(sheet.cell(row=row_number, column=col).value)

    job_raw = value("job_sequence")
    job_sequence: int | None
    if job_raw == "":
        job_sequence = None
    else:
        try:
            job_sequence = int(float(job_raw))
        except ValueError:
            job_sequence = None

    return ApplicationRecord(
        row_number=row_number,
        date=value("date"),
        url=value("url"),
        website=value("website"),
        job_sequence=job_sequence,
        company=value("company"),
        position=value("position"),
        resume=value("resume"),
        cover_letter=value("cover_letter"),
        updated=value("updated"),
    )


def _find_row_index(
    sheet,
    columns: dict[str, int],
    *,
    job_sequence: int | None,
) -> int | None:
    """Return 1-based Excel row number to update, or None."""
    company_col = columns.get("company")
    sequence_col = columns.get("job_sequence")
    max_row = sheet.max_row

    if job_sequence is not None and sequence_col is not None:
        for row in range(max_row, 1, -1):
            raw = sheet.cell(row=row, column=sequence_col).value
            if raw is None or _cell_str(raw) == "":
                continue
            try:
                if int(float(raw)) == job_sequence:
                    return row
            except (TypeError, ValueError):
                continue

    if company_col is not None:
        for row in range(max_row, 1, -1):
            if _cell_str(sheet.cell(row=row, column=company_col).value) == "":
                return row

    return None


def record_job_start(url: str, job_sequence: int) -> int:
    """
    Append a new job-start row. Returns 1-based Excel row number.
    """
    link = url.strip()
    if not link:
        raise ValueError("Job posting URL must be non-empty after trim")

    with _lock:
        ensure_workbook()
        path = tracker_path()
        workbook = load_workbook(path)
        try:
            sheet = workbook.active
            columns = _header_map(sheet)
            row_number = max(sheet.max_row, 1) + 1
            if sheet.max_row == 1 and _cell_str(sheet.cell(1, 1).value) == "":
                row_number = 2

            now = datetime.now()
            values = {
                "date": date.today().isoformat(),
                "url": link,
                "website": _website_from_url(link),
                "job_sequence": job_sequence,
                "company": "",
                "position": "",
                "resume": "N",
                "cover_letter": "N",
                "updated": now,
            }
            for name, value in values.items():
                col = columns.get(name)
                if col is not None:
                    sheet.cell(row=row_number, column=col, value=value)

            workbook.save(path)
            return row_number
        finally:
            workbook.close()


def record_resume_generated(
    company: str,
    position: str,
    *,
    job_sequence: int | None = None,
) -> None:
    """Update an existing row with resume generated (does not create a row)."""
    _record_document_generated(
        company,
        position,
        job_sequence=job_sequence,
        resume=True,
    )


def record_cover_letter_generated(
    company: str,
    position: str,
    *,
    job_sequence: int | None = None,
) -> None:
    """Update an existing row with cover letter generated (does not create a row)."""
    _record_document_generated(
        company,
        position,
        job_sequence=job_sequence,
        cover_letter=True,
    )


def _record_document_generated(
    company: str,
    position: str,
    *,
    job_sequence: int | None,
    resume: bool = False,
    cover_letter: bool = False,
) -> None:
    org = company.strip()
    role = position.strip()
    if not org or not role:
        raise ValueError("company and position must be non-empty after trim")

    with _lock:
        ensure_workbook()
        path = tracker_path()
        workbook = load_workbook(path)
        try:
            sheet = workbook.active
            columns = _header_map(sheet)
            row_number = _find_row_index(sheet, columns, job_sequence=job_sequence)
            if row_number is None:
                raise ValueError(
                    "No matching application row found "
                    "(job_sequence or pending Start row)"
                )

            now = datetime.now()
            if "company" in columns:
                sheet.cell(row=row_number, column=columns["company"], value=org)
            if "position" in columns:
                sheet.cell(row=row_number, column=columns["position"], value=role)
            if resume and "resume" in columns:
                sheet.cell(row=row_number, column=columns["resume"], value="Y")
            if cover_letter and "cover_letter" in columns:
                sheet.cell(row=row_number, column=columns["cover_letter"], value="Y")
            if "updated" in columns:
                sheet.cell(row=row_number, column=columns["updated"], value=now)

            workbook.save(path)
        finally:
            workbook.close()


def load_application_log(*, base: Path | None = None) -> list[ApplicationRecord]:
    """Return all data rows from the tracker workbook."""
    with _lock:
        path = tracker_path(base=base)
        if not path.is_file():
            ensure_workbook(base=base)
            return []

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            columns = _header_map(sheet)
            records: list[ApplicationRecord] = []
            for row in range(2, sheet.max_row + 1):
                if all(
                    _cell_str(sheet.cell(row=row, column=col).value) == ""
                    for col in range(1, len(HEADERS) + 1)
                ):
                    continue
                records.append(_row_to_record(sheet, row, columns))
            return records
        finally:
            workbook.close()


def application_stats(*, base: Path | None = None) -> dict[str, int]:
    records = load_application_log(base=base)
    companies = {r.company for r in records if r.company}
    return {
        "total_applications": len(records),
        "unique_companies": len(companies),
        "with_resume": sum(1 for r in records if r.resume.upper() == "Y"),
        "with_cover_letter": sum(1 for r in records if r.cover_letter.upper() == "Y"),
    }


def try_record_job_start(url: str, job_sequence: int) -> None:
    """Non-fatal wrapper for GUI / callers."""
    try:
        row = record_job_start(url, job_sequence)
        _log.info("Application tracker: job start recorded (row %s)", row)
    except Exception as exc:
        _log.warning("Application tracker: job start not recorded: %s", exc)


def try_record_resume_generated(
    company: str,
    position: str,
    *,
    job_sequence: int | None = None,
) -> None:
    try:
        record_resume_generated(company, position, job_sequence=job_sequence)
        _log.info("Application tracker: resume recorded for %s / %s", company, position)
    except Exception as exc:
        _log.warning("Application tracker: resume not recorded: %s", exc)


def try_record_cover_letter_generated(
    company: str,
    position: str,
    *,
    job_sequence: int | None = None,
) -> None:
    try:
        record_cover_letter_generated(company, position, job_sequence=job_sequence)
        _log.info(
            "Application tracker: cover letter recorded for %s / %s",
            company,
            position,
        )
    except Exception as exc:
        _log.warning("Application tracker: cover letter not recorded: %s", exc)
